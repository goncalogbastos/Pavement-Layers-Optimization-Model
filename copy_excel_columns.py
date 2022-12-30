import openpyxl


def copy_excel_columns(source_file, target_file, source_columns, target_columns):

    source_wb = openpyxl.load_workbook(source_file)
    target_wb = openpyxl.load_workbook(target_file)

    source_sheet_names = source_wb.sheetnames
    target_sheet_names = target_wb.sheetnames

    for i, sheet_name in enumerate(source_sheet_names):

        source_sheet = source_wb[sheet_name]
        target_sheet = target_wb[target_sheet_names[i]]

        for row in source_sheet.iter_rows():

            for j, cell in enumerate(row):

                if j in source_columns:

                    target_cell = target_sheet.cell(
                        row=cell.row, column=target_columns[source_columns.index(j)]
                    )
                    target_cell.value = None

                    target_cell.value = cell.value
                    cell.value = None

    target_wb.save(target_file)



print("\n############################################")
print('\nGetting data ready for analysis...')
print("\n############################################")


copy_excel_columns(
    "./output/layerSolutions_left.xlsx",
    "./src/layerSolutions_left_Civil.xlsx",
    [1, 2, 3, 4],
    [3, 4, 5, 6],
)
copy_excel_columns(
    "./output/layerSolutions_left.xlsx",
    "./src/layerSolutions_left_Civil.xlsx",
    [5, 6, 7, 8],
    [8, 9, 10, 11],
)
copy_excel_columns(
    "./output/layerSolutions_left.xlsx",
    "./src/layerSolutions_left_Civil.xlsx",
    [9, 10, 11, 12],
    [13, 14, 15, 16],
)
copy_excel_columns(
    "./output/layerSolutions_left.xlsx",
    "./src/layerSolutions_left_Civil.xlsx",
    [13, 14, 15, 16],
    [18, 19, 20, 21],
)

copy_excel_columns(
    "./output/layerSolutions_right.xlsx",
    "./src/layerSolutions_right_Civil.xlsx",
    [1, 2, 3, 4],
    [3, 4, 5, 6],
)
copy_excel_columns(
    "./output/layerSolutions_right.xlsx",
    "./src/layerSolutions_right_Civil.xlsx",
    [5, 6, 7, 8],
    [8, 9, 10, 11],
)
copy_excel_columns(
    "./output/layerSolutions_right.xlsx",
    "./src/layerSolutions_right_Civil.xlsx",
    [9, 10, 11, 12],
    [13, 14, 15, 16],
)
copy_excel_columns(
    "./output/layerSolutions_right.xlsx",
    "./src/layerSolutions_right_Civil.xlsx",
    [13, 14, 15, 16],
    [18, 19, 20, 21],
)
